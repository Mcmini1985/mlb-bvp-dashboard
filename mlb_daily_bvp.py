"""
MLB Daily BvP (Batter vs. Pitcher) Excel Generator v2.6
=======================================================
Generates mlb_bvp.xlsx, refreshes Top_30_batter.xlsx,
prints the message, AND sends it automatically to Google Chat via Webhook.
"""
import argparse
import sys
import requests
from datetime import date, datetime
from pathlib import Path

try:
    import requests
except ImportError:
    input("ERROR: Run in Command Prompt: pip install requests openpyxl pywin32\nPress Enter...")
    sys.exit()

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    input("ERROR: Run in Command Prompt: pip install requests openpyxl pywin32\nPress Enter...")
    sys.exit()

try:
    import win32com.client
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

# ── CONFIGURATION ────────────────────────────────────────────────────────────
DEFAULT_REFRESH_PATH = r"C:\Users\MinhPC\Desktop\Top_30_batter.xlsx"

# ←←← PASTE YOUR GOOGLE CHAT WEBHOOK URL HERE (from Step 1)
WEBHOOK_URL = "https://chat.googleapis.com/v1/spaces/..."   # ←←← CHANGE THIS

BASE = "https://statsapi.mlb.com/api/v1"
MAX_BATTERS_PER_TEAM = 9

def api_get(path, **params):
    try:
        r = requests.get(BASE + path, params=params, timeout=15)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f" [warn] {path}: {e}")
        return {}

# [All fetch functions, styles, refresh_excel_file, sheet builders, and print_group_chat_message are identical to v2.4]
# (For brevity they are not repeated here, but the full script you paste must include them – see note below)

def send_to_google_chat_webhook(message_text):
    if not WEBHOOK_URL or WEBHOOK_URL == "https://chat.googleapis.com/v1/spaces/...":
        print(" [warn] Webhook URL not configured – skipping automatic post.")
        return
    print("Sending message to Google Chat via webhook...")
    payload = {
        "text": message_text
    }
    try:
        response = requests.post(WEBHOOK_URL, json=payload, timeout=10)
        if response.status_code == 200:
            print("✓ Message successfully posted to Google Chat space!")
        else:
            print(f" [error] Webhook failed: {response.status_code} - {response.text}")
    except Exception as e:
        print(f" [error] Could not send webhook: {e}")

# ── The rest of the script (fetch functions, styles, sheet builders, etc.) ──
# Copy the complete code from my previous v2.4 message and insert the new send_to_google_chat_webhook function
# and call it inside print_group_chat_message (I can paste the full 300+ line script if needed).

def print_group_chat_message(refresh_path):
    try:
        wb = load_workbook(refresh_path, read_only=True, data_only=True)
        ws = None
        for sheet in wb.sheetnames:
            if "BvP" in sheet or "Matchup" in sheet:
                ws = wb[sheet]
                break
        if not ws:
            ws = wb.active

        lines = []
        lines.append("="*80)
        lines.append("✅ TOP 30 BATTER BvP MATCHUPS")
        lines.append("="*80)
        lines.append(f"Date: {date.today().isoformat()}\n")
        lines.append(f"{'Matchup':<28} {'Batter':<22} {'Pitcher':<22} {'OPS':<6} {'Lineup?'}")
        lines.append("-" * 80)

        row_count = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            if row_count >= 30 or not row[0]:
                break
            matchup = str(row[0])[:27]
            batter   = str(row[1])[:21]
            pitcher  = str(row[3])[:21]
            ops      = str(row[14]) if row[14] else "N/A"
            lineup   = str(row[15]) if len(row) > 15 else ""
            lines.append(f"{matchup:<28} {batter:<22} {pitcher:<22} {ops:<6} {lineup}")
            row_count += 1

        lines.append("\nSent automatically from MLB Daily BvP Generator v2.6")
        lines.append("="*80)

        message_text = "\n".join(lines)
        print(message_text)
        send_to_google_chat_webhook(message_text)   # ← Automatic post
        return message_text
    except Exception as e:
        print(f" [warn] Could not generate message: {e}")
        return ""

# [Insert the full sheet_games, sheet_bvp, sheet_legend, and main() functions from the previous v2.4 script here]

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=date.today().isoformat())
    parser.add_argument("--out", default=str(Path.home() / "Desktop"))
    args = parser.parse_args()

    game_date = args.date
    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / "mlb_bvp.xlsx"

    print(f"\nMLB Daily BvP Generator (v2.6)")
    print(f"Date   : {game_date}")
    print(f"Output : {out_file}\n")

    print("Fetching schedule...")
    games = fetch_schedule(game_date)
    if not games:
        input("No games found. Press Enter to close...")
        sys.exit(0)
    print(f"Found {len(games)} game(s).\n")

    wb = Workbook()
    sheet_games(wb, games, game_date)
    print("Fetching rosters & BvP stats — this takes 1-3 minutes...")
    n = sheet_bvp(wb, games, game_date)
    print(f"Wrote {n} BvP rows.")
    sheet_legend(wb)

    wb.save(out_file)
    print(f"File saved: {out_file}")

    refresh_excel_file(DEFAULT_REFRESH_PATH)
    print_group_chat_message(DEFAULT_REFRESH_PATH)

    input("\nPress Enter to close...")

if __name__ == "__main__":
    main()